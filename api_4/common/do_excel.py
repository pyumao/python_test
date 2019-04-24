# -*- coding: utf-8 -*-
# @Time    : 2019/3/11 20:31
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py
from openpyxl import load_workbook
from api_4.common.project_path import case_path
from api_4.common.read_config import ReadConfig

class DoExcel:
    '''该类完成测试数据的读取以及测试结果的写回'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name#Excel工作簿文件名或地址
        self.sheet_name=sheet_name#表单名


    def read_data(self,section):
        '''从Excel读取数据，有返回值'''

        caseid = ReadConfig().get(section,'caseid')
        wb=load_workbook(case_path)
        sheet=wb[self.sheet_name]
        tel=self.get_tel()  #获取手机号

        #唯一的要求是什么？每一行数据要在一起 {} []
        #如何把每一行的数据存到一个空间里面去？ []
        #开始读取数据
        if caseid == '1':
            test_data=[]
            for i in range(2,sheet.max_row+1):
                row_data={}
                row_data['CaseId']=sheet.cell(i,1).value
                row_data['Module']=sheet.cell(i,2).value
                row_data['URL']=sheet.cell(i,3).value
                row_data['Title']=sheet.cell(i,4).value
                row_data['Method']=sheet.cell(i,5).value
                if sheet.cell(i,6).value.find('tel')!=-1: #从列表中获取params值 是字符串 查找字符串中是否有tel字符没有返回-1 也可以用成员运算符in
                    row_data['Params'] = sheet.cell(i, 6).value.replace('tel',str(tel))
                    # print(type(tel))
                    # print(tel+1)
                    self.new_tel(1,2,tel+1)
                    # print(self.new_tel(1,2,tel+1))
                else:
                    row_data['Params'] = sheet.cell(i, 6).value
                row_data['SQL']=sheet.cell(i,7).value
                row_data['ExpectedResult']=sheet.cell(i,8).value
                test_data.append(row_data)
            wb.close()
            return test_data
        else:
            test_data = []
            for i in eval(caseid):
                row_data = {}
                row_data['CaseId'] = sheet.cell(i, 1).value
                row_data['Module'] = sheet.cell(i, 2).value
                row_data['URL'] = sheet.cell(i, 3).value
                row_data['Title'] = sheet.cell(i, 4).value
                row_data['Method'] = sheet.cell(i, 5).value
                if sheet.cell(i,6).value.find('tel')!=-1: #从列表中获取params值 是字符串 查找字符串中是否有tel字符没有返回-1 也可以用成员运算符in
                    row_data['Params'] = sheet.cell(i, 6).value.replace('tel',str(tel))
                    self.new_tel(1,2,tel+1)
                else:
                    row_data['Params'] = sheet.cell(i, 6).value
                row_data['SQL'] = sheet.cell(i, 7).value
                row_data['ExpectedResult'] = sheet.cell(i, 8).value
                test_data.append(row_data)
            wb.close()
            return test_data

    def get_tel(self):
        '''获取已存储的手机号'''
        wb=load_workbook(self.file_name)
        sheet=wb['set']
        tel=sheet.cell(1,2).value
        wb.close()
        return tel

    def new_tel(self,row,column,value):
        wb=load_workbook(self.file_name)
        sheet=wb['set']
        sheet.cell(1,2).value=value
        wb.save(case_path)
        wb.close()


    def write_back(self,row,column,value):
        '''写回测试结果到Excel中'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        sheet.cell(row,column).value=value

        wb.save(case_path)
        wb.close()

if __name__ == '__main__':
    file_name = case_path
    sheet_name = 'register'
    test_data = DoExcel(file_name, sheet_name).read_data('CaseRegister')
    print(type(test_data))

