# -*- codiing: utf-8 -*-
# @time      : 2019/3/23 0023 下午 2:13
# @Author    : yu
# @Email     : 
# @File      : learn_excel.py

#python的openpyxl模块来操作excel
#为什么要操作excel
#禅道 xmind  excel

#pip install  openpyxl
#openpyxl 里面 load_workbook是打开excel文件的
#work_book sheet cell
#学习的时候可以类比  open函数
#操作excel xlsx结尾的文件

from openpyxl import load_workbook
#打开工作簿
# wb=load_workbook('C:\Users\Administrator\python144\class_6\class_0222\python_14.xlsx')
wb=load_workbook('python_14.xlsx')
#
#定位表单
sheet=wb['Sheet1']

#定位单元格
# res=sheet.cell(row=1,column=3).value

# #确认最大行和列
# print(sheet.max_row)
# print(sheet.max_column)
#
#
# #读出某一行的数据
# s=[]
# for i in range(1,sheet.max_column+1):
#     res=sheet.cell(row=3,column=i).value
#     s.append(res)
# print(s)
# #
#读出所有行和列
s=[]

for i in range(1,sheet.max_row+1):
    s_1 = []
    for j in range(1,sheet.max_column+1):
        res=(sheet.cell(i,j).value)
        s_1=s.append(res)
    s.append(s_1)
print(s)

# #写数据
# sheet.cell(3,2,'zuozhe')  #行  列  value值
# #方法二
# sheet.cell(3,2).value='zuozhe'
#
# #只要进行了单元格的更改都要去进行保存  保存工作簿
#
# wb.save('python_14.xlsx')

#新建工作簿
# from openpyxl import workbook
#
# wb =workbook.Workbook()
#
# wb.save('python16.xlsx')