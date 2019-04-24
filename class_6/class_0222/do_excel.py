# -*- codiing: utf-8 -*-
# @time      : 2019/3/23 0023 下午 9:22
# @Author    : yu
# @Email     : 
# @File      : do_excel.py
from configparser import ConfigParser
from openpyxl import load_workbook
from openpyxl import workbook
from week_7.class_0225.read_config import ReadConfig
from week_7.class_0227.my_log import MyLog
logger=MyLog()
class DoExcel:

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def read_excel(self,button):
        '''读取excel中的数据，将没一行作为一个子列表，该行的每个单元格数据作为此子列表元素，'''
        wb = load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        if button == '1':
            test_data=[]  #所有的列表都存储在该列表中
            for i in range(2,sheet.max_row+1):
                row_data=[]  #每一行数据存储在一个列表中
                for j in range(1,sheet.max_column+1):
                    row_data.append(sheet.cell(i,j).value)
                test_data.append(row_data)
            return test_data
        else:
            test_data = []  # 所有的列表都存储在该列表中
            for i in  eval(button):
                row_data = []  # 每一行数据存储在一个列表中
                for j in range(1, sheet.max_column + 1):
                    row_data.append(sheet.cell(i+1, j).value)
                test_data.append(row_data)
            return test_data

        # test_data = []  # 所有的列表都存储在该列表中
        # for i in range(2, sheet.max_row + 1):
        #     row_data = {}  # 每一行数据存储在一个字典中
        #     row_data['CaseId']=sheet.cell(i,1).value
        #     row_data['Title'] = sheet.cell(i, 2).value
        #     row_data['Mudule'] = sheet.cell(i, 3).value
        #     row_data['TestData'] = sheet.cell(i, 4).value
        #     row_data['ExcepteResult'] = sheet.cell(i, 5).value
        #     row_data['ActualResult'] = sheet.cell(i, 6).value
        #     row_data['TestResult'] = sheet.cell(i, 7).value
        #     test_data.append(row_data)
        # return test_data
    #
    # def write_excel(self,row,column,value):
    #     '''在指定单元格中写数据，并存储到excel中'''
    #     wb=load_workbook(self.file_name)
    #     sheet=wb[self.sheet_name]
    #     logger.warning('开始写入数据')
    #     try:
    #         sheet.cell(row,column).value=value
    #         wb.save(self.file_name)
    #     except Exception as e:
    #         logger.error(e)
    #         wb.close()  #每次数据更改完后，关闭文件
    #     logger.warning('数据写入结束')
    # def create_excel(self):
    #     '''新建工作簿和表单'''
    #     wb=workbook.Workbook()
    #     wb.save(self.file_name)  # 创建新的工作簿
    #     wb.create_sheet(self.sheet_name) #创建表单


if __name__ == '__main__':
    button=ReadConfig('C:\\Users\Administrator\python144\week_7\class_0227\log.conf').get('CASE','button')
    T=DoExcel('python_14.xlsx','Sheet1')
    print(T.read_excel(button))
    # print(T.write_excel(1,7,'ok'))